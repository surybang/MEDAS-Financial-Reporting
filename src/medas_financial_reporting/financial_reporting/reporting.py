"""Reporting generation functions."""

from pathlib import Path

from loguru import logger
from openpyxl import load_workbook

from medas_financial_reporting.config import (
    INDICATORS,
    LOCAL_OUTPUT,
    LOCAL_TEMPLATE,
    SHEET_INDICATORS,
)


def write_data_to_excel(df) -> None:
    """
    Insère le DataFrame dans la feuille DATA du template.

    Args:
        df: DataFrame nettoyé.
    """
    import pandas as pd

    logger.info("Insertion des données dans le template")
    try:
        with pd.ExcelWriter(
            LOCAL_TEMPLATE,
            mode="a",
            engine="openpyxl",
            if_sheet_exists="replace",
        ) as writer:
            df.to_excel(writer, sheet_name="DATA", index=False)
        logger.success("Données insérées")
    except Exception as e:
        logger.error(f"Impossible d'insérer les données : {e}")
        raise RuntimeError(f"Impossible d'insérer les données : {e}") from e


def fill_indicators(
    input_path: Path | str = LOCAL_TEMPLATE,
    output_path: Path | str = LOCAL_OUTPUT,
    data_sheet: str = "DATA",
) -> None:
    """
    Remplit les indicateurs dans la feuille Indicateurs.

    Args:
        input_path: Chemin vers le fichier Excel source.
        output_path: Chemin vers le fichier Excel de sortie.
        data_sheet: Nom de la feuille de données.
    """
    logger.info("Remplissage des indicateurs")

    def formula_countif(col: str, val: str) -> str:
        return f'=COUNTIF({data_sheet}!{col}:{col}, "{val}")'

    def formula_countifs(pairs: list[tuple[str, str]]) -> str:
        conditions = ", ".join(
            f'{data_sheet}!{col}:{col}, "{val}"' for col, val in pairs
        )
        return f"=COUNTIFS({conditions})"

    def formula_sum(range_str: str) -> str:
        return f"=SUM({range_str})"

    wb = load_workbook(input_path)
    ws = wb[SHEET_INDICATORS]

    for item in INDICATORS:
        cell = f"E{item['row']}"
        if item["formule"] == "COUNTIF":
            ws[cell] = formula_countif(*item["args"][0])
        elif item["formule"] == "COUNTIFS":
            ws[cell] = formula_countifs(item["args"])
        elif item["formule"] == "SUM":
            ws[cell] = formula_sum(item["args"])
        else:
            raise ValueError(f"Formule inconnue : {item['formule']}")

    wb.save(output_path)
    wb.close()
    logger.success(f"Reporting généré : {output_path}")
