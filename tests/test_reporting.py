"""Tests for reporting.py"""

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from medas_financial_reporting.financial_reporting.reporting import (
    fill_indicators,
    write_data_to_excel,
)


@pytest.fixture
def minimal_workbook(tmp_path):
    """Workbook minimal avec les feuilles DATA et Indicateurs."""
    path = tmp_path / "test_reporting.xlsx"
    wb = Workbook()
    wb.active.title = "DATA"
    wb.create_sheet("Indicateurs")
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def sample_df():
    """DataFrame minimal simulant les données nettoyées."""
    return pd.DataFrame(
        {
            "type_client": ["PP", "PM"],
            "score": ["V", "S"],
            "score_prev": ["N", "O"],
            "id_agent": ["AUTO", "MANUEL"],
            "drc_complet": [True, False],
        }
    )


class TestFillIndicators:
    @pytest.mark.unit
    def test_countif_formula(self, minimal_workbook):
        fill_indicators(input_path=minimal_workbook, output_path=minimal_workbook)
        wb = load_workbook(minimal_workbook, data_only=False)
        ws = wb["Indicateurs"]
        assert ws["E8"].value == '=COUNTIF(DATA!B:B, "PP")'
        wb.close()

    @pytest.mark.unit
    def test_sum_formula(self, minimal_workbook):
        """Vérifie que la formule SUM est correctement générée."""
        fill_indicators(input_path=minimal_workbook, output_path=minimal_workbook)
        wb = load_workbook(minimal_workbook, data_only=False)
        ws = wb["Indicateurs"]
        assert ws["E10"].value == "=SUM(E8:E9)"
        wb.close()

    @pytest.mark.unit
    def test_countifs_formula(self, minimal_workbook):
        """Vérifie que la formule COUNTIFS est correctement générée."""
        fill_indicators(input_path=minimal_workbook, output_path=minimal_workbook)
        wb = load_workbook(minimal_workbook, data_only=False)
        ws = wb["Indicateurs"]
        assert ws["E14"].value == '=COUNTIFS(DATA!B:B, "PP", DATA!D:D, "V")'
        wb.close()

    @pytest.mark.unit
    def test_unknown_formula_raises(self, minimal_workbook):
        """Une formule inconnue lève une ValueError."""
        bad_indicators = [{"row": 99, "formule": "UNKNOWN", "args": []}]
        with pytest.raises(ValueError, match="Formule inconnue"):
            fill_indicators(
                input_path=minimal_workbook,
                output_path=minimal_workbook,
                indicators=bad_indicators,
            )


class TestWriteDataToExcel:
    @pytest.mark.unit
    def test_data_inserted(self, minimal_workbook, sample_df):
        """Vérifie que les données sont bien insérées dans la feuille DATA."""
        write_data_to_excel(sample_df, input_path=minimal_workbook)
        wb = load_workbook(minimal_workbook, data_only=False)
        ws = wb["DATA"]
        assert ws.cell(row=1, column=1).value == "type_client"
        assert ws.cell(row=2, column=1).value == "PP"
        wb.close()
