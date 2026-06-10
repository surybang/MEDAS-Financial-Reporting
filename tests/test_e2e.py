"""Tests End-to-End."""

import pytest
from openpyxl import load_workbook

from medas_financial_reporting.config import LOCAL_OUTPUT, S3_BUCKET, S3_OUTPUT_KEY
from medas_financial_reporting.financial_reporting.main import main
from medas_financial_reporting.storage import get_fs


@pytest.fixture(scope="session", autouse=True)
def run_pipeline():
    """Lance le pipeline une seule fois pour tous les tests E2E."""
    main()


@pytest.mark.e2e
def test_reporting_local_exists():
    """Vérifie que le reporting a bien été généré localement."""
    assert LOCAL_OUTPUT.exists()


@pytest.mark.e2e
def test_reporting_uploaded_to_minio():
    """Vérifie que le reporting est bien présent sur MinIO."""
    fs = get_fs()
    assert fs.exists(f"{S3_BUCKET}/{S3_OUTPUT_KEY}")


@pytest.mark.e2e
def test_reporting_contains_expected_sheets():
    """Vérifie que le reporting contient les feuilles attendues."""
    wb = load_workbook(LOCAL_OUTPUT)
    assert "DATA" in wb.sheetnames
    assert "Indicateurs" in wb.sheetnames
    wb.close()


@pytest.mark.e2e
def test_indicators_are_filled():
    """Vérifie que les indicateurs ne sont pas vides."""
    wb = load_workbook(LOCAL_OUTPUT)
    ws = wb["Indicateurs"]
    assert ws["E8"].value is not None
    assert ws["E10"].value is not None
    wb.close()
